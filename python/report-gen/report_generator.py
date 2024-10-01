import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from sqlalchemy import create_engine
import re

class ReportGenerator:
    def __init__(self, config, password, username, hostname, database):
        self.config = config
        self.current_date = datetime.now().strftime("%Y-%m-%d")
        self.password = password
        self.username = username
        self.hostname = hostname
        self.database = database
        self.engine = self.connect_to_database()

    def connect_to_database(self):
        return create_engine(f"mysql+mysqlconnector://{self.username}:{self.password}@{self.hostname}/{self.database}")

    def fetch_data(self):
        selected_columns = []
        for column, settings in self.config['Columns'].items():
            alias_name = settings.get('aliasName', '')
            if alias_name:
                if ' ' in alias_name:
                    alias_name = f"'{alias_name}'"
                selected_columns.append(f"{column} AS {alias_name}")
            else:
                selected_columns.append(column)

        selected_columns_str = ", ".join(selected_columns)
        
        query = f"""
            SELECT {selected_columns_str} FROM daily_trans dt 
            INNER JOIN daily_collection dc 
            ON dt.trans_id=dc.trans_id 
            WHERE dt.status='Y' 
            AND dc.status='Y' 
            AND dc.multi_rec='N' 
            AND coll_date between '2024-09-01' and '2024-09-02'
     """
        data = pd.read_sql(query, self.engine)
        data.sort_values(by=self.config['OrderBy'], inplace=True)
        return data

    def apply_conditions(self, data):
        # Create a mapping from original column names to their alias names
        column_alias_map = {column: settings['aliasName'] for column, settings in self.config['Columns'].items() if 'aliasName' in settings}

        # Apply default values and conditions from the config for each column
        for column, settings in self.config['Columns'].items():
            actual_column = column_alias_map.get(column, column)

            if actual_column in data.columns:
                condition = settings.get('condition', [])
                if isinstance(condition, list) and condition:  # Check if conditions exist
                    for cond in condition:
                        if actual_column in cond and cond[actual_column] == 'ALL':
                            default_value = cond.get('defaultValue')
                            if default_value is not None:
                                data[actual_column] = default_value
                        else:
                            for field, value in cond.items():
                                # Make sure to map field names to their aliases if they exist
                                actual_field = column_alias_map.get(field, field)
                                if actual_field in data.columns:
                                    data[actual_column] = data[actual_column].replace(value, cond.get("defaultValue"))

                if 'defaultValue' in settings and settings['defaultValue'] is not None:
                    default_value = settings['defaultValue']
                    data[actual_column] = data[actual_column].fillna(default_value)
                    data[actual_column] = data[actual_column].replace('', default_value)

        return data

    def create_custom_columns(self, data):
        # Create a dictionary to hold conditions for each custom column
        custom_col_data = {}
        
        for custom_col, settings in self.config.get('includeCustomColumns', {}).items():
            if not settings.get('enabled', "False"):
                continue  # Skip if the custom column is not enabled

            # Initialize the custom column with default value
            custom_col_data[custom_col] = settings.get('defaultValue', None)
            conditions = settings.get('condition', [])

            if isinstance(conditions, list) and conditions:
                for condition in conditions:
                    column_name = condition.get('column')
                    match_type = condition.get('type', 'string')
                    action = condition.get('action')
                    match_value = condition.get('matchValue')
                    replace_value = condition.get('ReplaceValue')

                    if column_name in data.columns:
                        # Convert match_value to the correct type if necessary
                        if match_type == 'int':
                            match_value = int(match_value)

                        # Prepare boolean masks for conditions
                        if action == '>':
                            mask = data[column_name] > match_value
                        elif action == '<':
                            mask = data[column_name] < match_value
                        elif action == '=':
                            if match_type == 'int':
                                mask = data[column_name] == match_value
                            elif match_type == 'string':
                                mask = data[column_name].astype(str) == match_value

                        # Apply the mask to set the replace_value
                        data.loc[mask, custom_col] = replace_value

            if custom_col in data.columns and data[custom_col].isnull().all() and settings.get('defaultValue') is not None:
                data[custom_col] = settings['defaultValue']

        # After processing all custom columns, insert them into the DataFrame
        for custom_col in custom_col_data.keys():
            if custom_col not in data.columns:
                # Determine the position to insert the new column
                position = settings.get('position')
                insert_index = data.columns.get_loc(position) + 1 if position in data.columns else len(data.columns)
                data.insert(insert_index, custom_col, custom_col_data[custom_col])

        return data

    def handle_bodmas(self, data):
        for custom_col, settings in self.config.get('includeCustomColumns', {}).items():
            if settings.get('enabled', "False"):
                conditions = settings.get('condition', [])
                if conditions:
                    for condition in conditions:
                        operation = condition.get('operation', None)
                        if operation:
                            for col in condition.get('Actioncolumns', "").split(','):
                                operation = operation.replace(col.strip(), f"data['{col.strip()}']")
                            try:
                                data[custom_col] = eval(operation)
                            except Exception as e:
                                print(f"Error evaluating operation '{operation}': {e}")

        return data

    def rename_columns(self, data):
        rename_dict = {column: settings['aliasName'] for column, settings in self.config['Columns'].items() if 'aliasName' in settings}
        data.rename(columns=rename_dict, inplace=True)
        return data

    def calculate_sums(self, data):
        sum_columns = [settings['aliasName'] for column, settings in self.config['Columns'].items() if 'sumBy' in settings and settings['sumBy'] == "True"]
        sum_columns += [custom_col for custom_col, settings in self.config.get('includeCustomColumns', {}).items() if settings.get('sumBy', "False") == "True"]
        data = data.loc[:, ~data.columns.duplicated()]
        return data, sum_columns

    def generate_excel(self, data, sum_columns):
        output_file = f"{self.config['ClientName'].replace(' ', '-')}-{self.current_date}.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if self.config['Sheet'] == 'Single':
                self._generate_single_sheet(data, sum_columns, writer)
            elif self.config['Sheet'] == 'Multi':
                self._generate_multi_sheet(data, sum_columns, writer)

        print(f"Report generated: {output_file}")
        
    def _generate_single_sheet(self, data, sum_columns, writer):
        grouped_data = data.groupby(self.config['OrderBy'], sort=False)
        total_columns = len(data.columns)
        ws = writer.book.create_sheet(title=self.current_date)

        # Merge header cells
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_columns)
        ws["A1"] = f"{self.config['OrgName']}"
        ws["A2"] = f"{self.config['ClientName']}"
        ws["A3"] = f"{self.current_date}"

        title_fill = PatternFill(start_color=self.config['TitleColour'], end_color=self.config['TitleColour'], fill_type="solid")
        for row in [1, 2, 3]:
            cell = ws[f"A{row}"]
            cell.fill = title_fill
            cell.font = Font(bold=True, color=self.config['TitleFontColour'])
            cell.alignment = Alignment(horizontal='center', vertical='center')

        headers = list(data.columns)
        for col_num, header in enumerate(headers, start=1):
            header_cell = ws.cell(row=4, column=col_num, value=header)
            header_cell.fill = PatternFill(start_color=self.config['HeaderColour'], end_color=self.config['HeaderColour'], fill_type="solid")
            header_cell.font = Font(bold=True, color=self.config['HeaderFontColour'])

        data_row = 5

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        solid_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type="solid")

        if self.config['TotalType'] == "grand":
            for row_num, row_data in data.iterrows():
                for col_num, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=data_row, column=col_num, value=value)
                    cell.border = thin_border
                data_row += 1

            grand_total_row = data_row

            # Insert individual grand total label if specified
            if self.config.get('IncludeIndividualGrandTotal', "False").lower() == "true":
                # Find the position for the label
                label_column_name = self.config.get("SumLabelPostionColumnName", None)
                if label_column_name and label_column_name in data.columns:
                    label_column_index = data.columns.get_loc(label_column_name) + 1
                else:
                    label_column_index = 1  # Default to the first column

                # Insert individual grand total label
                grand_individual_total_label = "GRAND TOTAL"
                ws.cell(row=grand_total_row, column=label_column_index, value=grand_individual_total_label).font = Font(bold=True)

            for sum_col in sum_columns:
                sum_column_index = list(data.columns).index(sum_col) + 1
                sum_column_letter = get_column_letter(sum_column_index)

                # Insert Excel sum formula for each sum column
                sum_formula = f"=SUM({sum_column_letter}5:{sum_column_letter}{grand_total_row - 1})"
                total_cell = ws[f"{sum_column_letter}{grand_total_row}"]
                total_cell.value = sum_formula
                total_cell.font = Font(bold=True)

                # Apply solid background to the entire sum row and thin borders
                for col_num in range(1, total_columns + 1):
                    sum_cell = ws.cell(row=grand_total_row, column=col_num)
                    sum_cell.fill = solid_fill
                    sum_cell.border = thin_border

        elif self.config['TotalType'] == "individual":
            all_group_rows = []  # To keep track of data rows without individual totals
            
            for key, group in grouped_data:
                group_start_row = data_row  # Track where the group starts
                all_group_rows.append((group_start_row, group_start_row + len(group) - 1))  # Store the range of data rows for each group
                
                # Write group data
                for row_num, row_data in group.iterrows():
                    for col_num, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=data_row, column=col_num, value=value)
                        cell.border = thin_border
                    data_row += 1

                # Write individual total for the group (focus on labels and formulas)
                individual_total_row = data_row

                # Convert label to uppercase if needed
                label = f"{key}".upper()

                # Find the column index for the label position using the new config key
                label_column_name = self.config.get("SumLabelPostionColumnName", None)  # None if not found
                if label_column_name and label_column_name in data.columns:
                    label_column_index = data.columns.get_loc(label_column_name) + 1  # +1 for 1-based index
                else:
                    label_column_index = 1  # Default to first column if not found

                # Insert the custom label for individual total
                ws.cell(row=individual_total_row, column=label_column_index, value=label).font = Font(bold=True)

                for sum_col in sum_columns:
                    sum_column_index = list(data.columns).index(sum_col) + 1
                    sum_column_letter = get_column_letter(sum_column_index)

                    # Insert Excel sum formula for individual totals
                    sum_formula = f"=SUM({sum_column_letter}{group_start_row}:{sum_column_letter}{individual_total_row - 1})"
                    total_cell = ws[f"{sum_column_letter}{individual_total_row}"]
                    total_cell.value = sum_formula
                    total_cell.font = Font(bold=True)

                    # Apply solid background to the entire sum row and thin borders
                    for col_num in range(1, total_columns + 1):
                        sum_cell = ws.cell(row=individual_total_row, column=col_num)
                        sum_cell.fill = solid_fill
                        sum_cell.border = thin_border
                        
                data_row += 1

            # Now calculate the grand total, but only if IncludeIndividualGrandTotal is true
            if self.config.get('IncludeIndividualGrandTotal', "False") == "True":
                grand_total_row = data_row
 
                # Find the column index for the label position using the new config key
                label_column_name = self.config.get("SumLabelPostionColumnName", None)  # None if not found
                if label_column_name and label_column_name in data.columns:
                    label_column_index = data.columns.get_loc(label_column_name) + 1  # +1 for 1-based index
                else:
                    label_column_index = 1

                # Insert the custom label
                ws.cell(row=grand_total_row, column=label_column_index, value="INDIVIDUAL GRAND TOTAL").font = Font(bold=True)

                for sum_col in sum_columns:
                    sum_column_index = list(data.columns).index(sum_col) + 1
                    sum_column_letter = get_column_letter(sum_column_index)
                    
                    # Create the SUM formula excluding the individual total rows
                    sum_ranges = [f"{sum_column_letter}{start}:{sum_column_letter}{end}" for start, end in all_group_rows]
                    sum_formula = f"=SUM({','.join(sum_ranges)})"
                    
                    # Apply the formula to the grand total row
                    total_cell = ws[f"{sum_column_letter}{grand_total_row}"]
                    total_cell.value = sum_formula
                    total_cell.font = Font(bold=True)

                    # Apply solid background to the entire sum row and thin borders
                    for col_num in range(1, total_columns + 1):
                        sum_cell = ws.cell(row=grand_total_row, column=col_num)
                        sum_cell.fill = solid_fill
                        sum_cell.border = thin_border

    def _generate_multi_sheet(self, data, sum_columns, writer):
        # Define border styles at the beginning of the method
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # Group data by the specified 'OrderBy' key
        grouped_data = data.groupby(self.config['OrderBy'], sort=False)

        for key, group in grouped_data:
            # Create a new sheet for each group
            ws = writer.book.create_sheet(title=str(key))
            total_columns = len(group.columns)

            # Merge header cells across the total number of columns
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_columns)

            # Set organization, client name, and current date in the merged header cells
            ws["A1"] = f"{self.config['OrgName']}"
            ws["A2"] = f"{self.config['ClientName']}"
            ws["A3"] = f"{self.current_date}"

            # Apply header styles (fill, font, alignment)
            title_fill = PatternFill(start_color=self.config['TitleColour'], end_color=self.config['TitleColour'], fill_type="solid")
            for row in [1, 2, 3]:
                cell = ws[f"A{row}"]
                cell.fill = title_fill
                cell.font = Font(bold=True, color=self.config['TitleFontColour'])
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Set headers for each column in row 4
            headers = list(group.columns)
            for col_num, header in enumerate(headers, start=1):
                header_cell = ws.cell(row=4, column=col_num, value=header)
                header_cell.fill = PatternFill(start_color=self.config['HeaderColour'], end_color=self.config['HeaderColour'], fill_type="solid")
                header_cell.font = Font(bold=True, color=self.config['HeaderFontColour'])
                header_cell.border = thin_border

            data_row = 5
            solid_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type="solid")

            # Fill data and apply borders to each data row
            for row_num, row_data in group.iterrows():
                for col_num, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=data_row, column=col_num, value=value)
                    cell.border = thin_border
                data_row += 1

            # Handle grand total row after data
            grand_total_row = data_row

            # Insert individual grand total label if specified
            if self.config.get('IncludeIndividualGrandTotal', "False").lower() == "true":
                label_column_name = self.config.get("SumLabelPostionColumnName", None)
                if label_column_name and label_column_name in group.columns:
                    label_column_index = group.columns.get_loc(label_column_name) + 1
                else:
                    label_column_index = 1

                # Insert individual grand total label
                grand_individual_total_label = "INDIVIDUAL GRAND TOTAL"
                ws.cell(row=grand_total_row, column=label_column_index, value=grand_individual_total_label).font = Font(bold=True)

            for sum_col in sum_columns:
                sum_column_index = list(group.columns).index(sum_col) + 1
                sum_column_letter = get_column_letter(sum_column_index)

                # Insert SUM formula for the column
                total_cell = ws[f"{sum_column_letter}{grand_total_row}"]
                start_row = 5
                total_cell.value = f"=SUM({sum_column_letter}{start_row}:{sum_column_letter}{grand_total_row-1})"
                total_cell.font = Font(bold=True)

                # Apply solid background to the entire sum row and thin borders
                for col_num in range(1, total_columns + 1):
                    sum_cell = ws.cell(row=grand_total_row, column=col_num)
                    sum_cell.fill = solid_fill
                    sum_cell.border = thin_border

            # Apply thin borders to all data rows
            for row in range(5, data_row):
                for col in range(1, total_columns + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

    def run_report(self):
        data = self.fetch_data()
        data = self.apply_conditions(data)
        data = self.rename_columns(data)
        data = self.create_custom_columns(data)
        data = self.handle_bodmas(data)
        data, sum_columns = self.calculate_sums(data)
        self.generate_excel(data, sum_columns)
